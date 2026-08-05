"""
M10-E03 Phase 3 - Campaign Ranked Candidate Export.

Mirrors test_campaign_candidate_scorecard_integration.py's real-workbook
round-trip technique (drain the StreamingResponse body, parse it with
openpyxl) for the integration-style tests, plus MagicMock-based tests for
filter pass-through, audit content, and edge cases - the same split every
other export test file in this codebase already uses.
"""
import asyncio
import io
from datetime import datetime, timezone
from types import SimpleNamespace
from unittest.mock import MagicMock
from uuid import uuid4

import openpyxl
import pytest

from app.enums.constants import ActionType, EntityType
from app.exceptions.campaign_exceptions import CampaignException
from app.models.pipeline import AIEvaluationStatus, AIRecommendation, PipelineStage
from app.services.campaign.campaign_candidate_service import CampaignCandidateService


def _drain_streaming_response_body(response) -> bytes:
    async def _collect():
        chunks = []
        async for chunk in response.body_iterator:
            chunks.append(chunk)
        return b"".join(chunks)

    return asyncio.run(_collect())


def _make_service(campaign_repo=None, campaign_candidate_repo=None, audit_service=None):
    return CampaignCandidateService(
        campaign_repo=campaign_repo or MagicMock(),
        campaign_candidate_repo=campaign_candidate_repo or MagicMock(),
        audit_service=audit_service or MagicMock(),
    )


def _make_campaign_candidate(
    composite_score=None, deterministic_score=None, semantic_score=None, effective_ai_score=None,
    pipeline_stage=PipelineStage.SCREENING, ai_recommendation=None,
    ai_evaluation_status=AIEvaluationStatus.PENDING, composite_score_computed_at=None,
):
    return SimpleNamespace(
        id=uuid4(), candidate_id=uuid4(), pipeline_stage=pipeline_stage,
        composite_score=composite_score, deterministic_score=deterministic_score,
        semantic_score=semantic_score, effective_ai_score=effective_ai_score,
        ai_recommendation=ai_recommendation, ai_evaluation_status=ai_evaluation_status,
        composite_score_computed_at=composite_score_computed_at,
    )


def _ranked_row(cc):
    """get_ranked_by_campaign returns (CampaignCandidate, Candidate, Resume) tuples."""
    return (cc, None, None)


# ----------------------------------------------------------------------
# Validation / not-found
# ----------------------------------------------------------------------

def test_raises_when_campaign_not_found():
    campaign_repo = MagicMock()
    campaign_repo.get_by_id.return_value = None
    service = _make_service(campaign_repo=campaign_repo)

    with pytest.raises(CampaignException) as exc_info:
        service.export_ranked_campaign_candidates(uuid4(), actor_id="hr-1", actor_role="HR_ADMIN")

    assert exc_info.value.status_code == 404


def test_rejects_composite_score_min_greater_than_max():
    campaign_repo = MagicMock()
    campaign_repo.get_by_id.return_value = SimpleNamespace(id=uuid4())
    service = _make_service(campaign_repo=campaign_repo)

    with pytest.raises(CampaignException) as exc_info:
        service.export_ranked_campaign_candidates(
            uuid4(), actor_id="hr-1", actor_role="HR_ADMIN",
            composite_score_min=80, composite_score_max=20,
        )

    assert exc_info.value.status_code == 422


# ----------------------------------------------------------------------
# Pagination is ignored - the complete filtered dataset is exported
# ----------------------------------------------------------------------

def test_ignores_pagination_and_fetches_the_entire_filtered_dataset():
    """
    The service must call get_ranked_by_campaign twice: once to learn the
    filtered total (page_size=1), then once more with page_size=total -
    never page=3/page_size=20 or any other UI-page-shaped call.
    """
    campaign = SimpleNamespace(id=uuid4())
    campaign_repo = MagicMock()
    campaign_repo.get_by_id.return_value = campaign
    ccs = [_make_campaign_candidate(composite_score=90 - i) for i in range(47)]
    campaign_candidate_repo = MagicMock()
    campaign_candidate_repo.get_ranked_by_campaign.side_effect = [
        ([], 47),
        ([_ranked_row(cc) for cc in ccs], 47),
    ]
    service = _make_service(campaign_repo=campaign_repo, campaign_candidate_repo=campaign_candidate_repo)

    service.export_ranked_campaign_candidates(campaign.id, actor_id="hr-1", actor_role="HR_ADMIN")

    assert campaign_candidate_repo.get_ranked_by_campaign.call_count == 2
    first_call, second_call = campaign_candidate_repo.get_ranked_by_campaign.call_args_list
    assert first_call.kwargs["page"] == 1
    assert first_call.kwargs["page_size"] == 1
    assert second_call.kwargs["page"] == 1
    assert second_call.kwargs["page_size"] == 47


def test_skips_second_call_when_campaign_has_zero_matching_candidates():
    campaign = SimpleNamespace(id=uuid4())
    campaign_repo = MagicMock()
    campaign_repo.get_by_id.return_value = campaign
    campaign_candidate_repo = MagicMock()
    campaign_candidate_repo.get_ranked_by_campaign.return_value = ([], 0)
    audit_service = MagicMock()
    service = _make_service(
        campaign_repo=campaign_repo, campaign_candidate_repo=campaign_candidate_repo, audit_service=audit_service,
    )

    response = service.export_ranked_campaign_candidates(campaign.id, actor_id="hr-1", actor_role="HR_ADMIN")

    assert campaign_candidate_repo.get_ranked_by_campaign.call_count == 1
    body = _drain_streaming_response_body(response)
    workbook = openpyxl.load_workbook(io.BytesIO(body))
    sheet = workbook.active
    assert sheet.max_row == 1  # header only
    audit_service.log.assert_called_once()
    assert audit_service.log.call_args.kwargs["details"]["rows_exported"] == 0


# ----------------------------------------------------------------------
# Filter/sort pass-through - reuses get_ranked_by_campaign exactly
# ----------------------------------------------------------------------

def test_passes_every_filter_through_to_get_ranked_by_campaign_unchanged():
    campaign = SimpleNamespace(id=uuid4())
    campaign_repo = MagicMock()
    campaign_repo.get_by_id.return_value = campaign
    campaign_candidate_repo = MagicMock()
    campaign_candidate_repo.get_ranked_by_campaign.return_value = ([], 0)
    service = _make_service(campaign_repo=campaign_repo, campaign_candidate_repo=campaign_candidate_repo)

    service.export_ranked_campaign_candidates(
        campaign.id, actor_id="hr-1", actor_role="HR_ADMIN",
        sort_by="semantic_score", sort_order="asc", pipeline_stage=PipelineStage.SHORTLISTED,
        composite_score_min=10, composite_score_max=90, ai_recommendation=AIRecommendation.SHORTLIST,
        ai_evaluation_status=AIEvaluationStatus.COMPLETED, include_pending=False, include_rejected=False,
        include_fraud=False, hr_override=True,
    )

    call = campaign_candidate_repo.get_ranked_by_campaign.call_args
    assert call.kwargs["sort_by"] == "semantic_score"
    assert call.kwargs["sort_order"] == "asc"
    assert call.kwargs["pipeline_stage"] == PipelineStage.SHORTLISTED
    assert call.kwargs["composite_score_min"] == 10
    assert call.kwargs["composite_score_max"] == 90
    assert call.kwargs["ai_recommendation"] == AIRecommendation.SHORTLIST
    assert call.kwargs["ai_evaluation_status"] == AIEvaluationStatus.COMPLETED
    assert call.kwargs["include_pending"] is False
    assert call.kwargs["include_rejected"] is False
    assert call.kwargs["include_fraud"] is False
    assert call.kwargs["hr_override"] is True


# ----------------------------------------------------------------------
# Audit - exactly one entry per export request
# ----------------------------------------------------------------------

def test_audit_logged_exactly_once_with_expected_shape():
    campaign = SimpleNamespace(id=uuid4())
    campaign_repo = MagicMock()
    campaign_repo.get_by_id.return_value = campaign
    cc = _make_campaign_candidate(composite_score=80.0)
    campaign_candidate_repo = MagicMock()
    campaign_candidate_repo.get_ranked_by_campaign.side_effect = [([], 1), ([_ranked_row(cc)], 1)]
    audit_service = MagicMock()
    service = _make_service(
        campaign_repo=campaign_repo, campaign_candidate_repo=campaign_candidate_repo, audit_service=audit_service,
    )

    service.export_ranked_campaign_candidates(
        campaign.id, actor_id="hr-1", actor_role="HR_ADMIN", pipeline_stage=PipelineStage.SHORTLISTED,
    )

    audit_service.log.assert_called_once()
    kwargs = audit_service.log.call_args.kwargs
    assert kwargs["action_type"] == ActionType.CANDIDATE_RANKING_EXPORTED
    assert kwargs["entity_type"] == EntityType.CAMPAIGN
    assert kwargs["entity_id"] == campaign.id
    assert kwargs["campaign_id"] == campaign.id
    assert kwargs["actor_id"] == "hr-1"
    assert kwargs["details"]["export_format"] == "XLSX"
    assert kwargs["details"]["rows_exported"] == 1
    assert kwargs["details"]["applied_filters"]["pipeline_stage"] == "SHORTLISTED"


# ----------------------------------------------------------------------
# Real XLSX round-trip: structure, content, and PII exclusion
# ----------------------------------------------------------------------

def test_export_produces_a_real_readable_workbook_with_expected_headers_and_no_pii():
    campaign = SimpleNamespace(id=uuid4())
    campaign_repo = MagicMock()
    campaign_repo.get_by_id.return_value = campaign
    computed_at = datetime.now(timezone.utc)
    cc = _make_campaign_candidate(
        composite_score=87.5, deterministic_score=90.0, semantic_score=0.8, effective_ai_score=75.0,
        pipeline_stage=PipelineStage.SHORTLISTED, ai_recommendation=AIRecommendation.SHORTLIST,
        composite_score_computed_at=computed_at,
    )
    campaign_candidate_repo = MagicMock()
    campaign_candidate_repo.get_ranked_by_campaign.side_effect = [([], 1), ([_ranked_row(cc)], 1)]
    service = _make_service(campaign_repo=campaign_repo, campaign_candidate_repo=campaign_candidate_repo)

    response = service.export_ranked_campaign_candidates(campaign.id, actor_id="hr-1", actor_role="HR_ADMIN")

    body = _drain_streaming_response_body(response)
    workbook = openpyxl.load_workbook(io.BytesIO(body))
    sheet = workbook.active

    header = [cell.value for cell in sheet[1]]
    assert header == [
        "Rank",
        "Candidate UUID",
        "Composite Score",
        "Deterministic Score",
        "Semantic Score",
        "AI Evaluation Score",
        "Pipeline Stage",
        "AI Recommendation",
        "Ranking Status",
        "Composite Score Computed At",
    ]
    for forbidden in ("Candidate Name", "Email", "Phone", "Resume"):
        assert forbidden not in header

    data_row = [cell.value for cell in sheet[2]]
    assert data_row[0] == 1  # rank
    assert data_row[1] == str(cc.candidate_id)
    assert data_row[2] == 87.5
    assert data_row[3] == 90.0
    assert data_row[4] == 0.8
    assert data_row[5] == 75.0
    assert data_row[6] == "SHORTLISTED"
    assert data_row[7] == "SHORTLIST"
    assert data_row[8] == "RANKED"
    assert data_row[9] == computed_at.strftime("%d-%b-%Y %I:%M %p")


def test_export_response_headers_declare_xlsx_attachment():
    campaign = SimpleNamespace(id=uuid4())
    campaign_repo = MagicMock()
    campaign_repo.get_by_id.return_value = campaign
    campaign_candidate_repo = MagicMock()
    campaign_candidate_repo.get_ranked_by_campaign.return_value = ([], 0)
    service = _make_service(campaign_repo=campaign_repo, campaign_candidate_repo=campaign_candidate_repo)

    response = service.export_ranked_campaign_candidates(campaign.id, actor_id="hr-1", actor_role="HR_ADMIN")

    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response.headers["content-disposition"].startswith('attachment; filename="candidate_ranking_export_')
    assert response.headers["content-disposition"].endswith('.xlsx"')


def test_rank_is_1_based_and_sequential_across_the_full_dataset():
    campaign = SimpleNamespace(id=uuid4())
    campaign_repo = MagicMock()
    campaign_repo.get_by_id.return_value = campaign
    ccs = [_make_campaign_candidate(composite_score=100 - i) for i in range(5)]
    campaign_candidate_repo = MagicMock()
    campaign_candidate_repo.get_ranked_by_campaign.side_effect = [
        ([], 5), ([_ranked_row(cc) for cc in ccs], 5),
    ]
    service = _make_service(campaign_repo=campaign_repo, campaign_candidate_repo=campaign_candidate_repo)

    response = service.export_ranked_campaign_candidates(campaign.id, actor_id="hr-1", actor_role="HR_ADMIN")

    body = _drain_streaming_response_body(response)
    workbook = openpyxl.load_workbook(io.BytesIO(body))
    sheet = workbook.active
    ranks = [row[0].value for row in sheet.iter_rows(min_row=2)]
    assert ranks == [1, 2, 3, 4, 5]


# ----------------------------------------------------------------------
# Edge cases: pending, rejected, fraud, missing composite, large dataset
# ----------------------------------------------------------------------

def test_all_pending_candidates_export_with_null_scores_and_pending_status():
    campaign = SimpleNamespace(id=uuid4())
    campaign_repo = MagicMock()
    campaign_repo.get_by_id.return_value = campaign
    cc = _make_campaign_candidate(composite_score=None, ai_evaluation_status=AIEvaluationStatus.PENDING)
    campaign_candidate_repo = MagicMock()
    campaign_candidate_repo.get_ranked_by_campaign.side_effect = [([], 1), ([_ranked_row(cc)], 1)]
    service = _make_service(campaign_repo=campaign_repo, campaign_candidate_repo=campaign_candidate_repo)

    response = service.export_ranked_campaign_candidates(campaign.id, actor_id="hr-1", actor_role="HR_ADMIN")

    body = _drain_streaming_response_body(response)
    sheet = openpyxl.load_workbook(io.BytesIO(body)).active
    data_row = [cell.value for cell in sheet[2]]
    assert data_row[2] is None  # composite score
    assert data_row[8] == "PENDING"


def test_all_fraud_flagged_candidates_export_correctly():
    campaign = SimpleNamespace(id=uuid4())
    campaign_repo = MagicMock()
    campaign_repo.get_by_id.return_value = campaign
    cc = _make_campaign_candidate(composite_score=55.0, pipeline_stage=PipelineStage.FRAUD_REVIEW)
    campaign_candidate_repo = MagicMock()
    campaign_candidate_repo.get_ranked_by_campaign.side_effect = [([], 1), ([_ranked_row(cc)], 1)]
    service = _make_service(campaign_repo=campaign_repo, campaign_candidate_repo=campaign_candidate_repo)

    response = service.export_ranked_campaign_candidates(
        campaign.id, actor_id="hr-1", actor_role="HR_ADMIN", include_fraud=True,
    )

    body = _drain_streaming_response_body(response)
    sheet = openpyxl.load_workbook(io.BytesIO(body)).active
    assert sheet[2][6].value == "FRAUD_REVIEW"


def test_large_dataset_exports_every_row_in_one_workbook():
    campaign = SimpleNamespace(id=uuid4())
    campaign_repo = MagicMock()
    campaign_repo.get_by_id.return_value = campaign
    ccs = [_make_campaign_candidate(composite_score=float(i % 100)) for i in range(1200)]
    campaign_candidate_repo = MagicMock()
    campaign_candidate_repo.get_ranked_by_campaign.side_effect = [
        ([], 1200), ([_ranked_row(cc) for cc in ccs], 1200),
    ]
    service = _make_service(campaign_repo=campaign_repo, campaign_candidate_repo=campaign_candidate_repo)

    response = service.export_ranked_campaign_candidates(campaign.id, actor_id="hr-1", actor_role="HR_ADMIN")

    body = _drain_streaming_response_body(response)
    sheet = openpyxl.load_workbook(io.BytesIO(body)).active
    assert sheet.max_row == 1201  # header + 1200 rows
    second_call = campaign_candidate_repo.get_ranked_by_campaign.call_args_list[1]
    assert second_call.kwargs["page_size"] == 1200


# ----------------------------------------------------------------------
# Backward compatibility - unrelated existing exports/methods unaffected
# ----------------------------------------------------------------------

def test_does_not_disturb_existing_get_ranked_campaign_candidates_method():
    """Phase 1's read-only ranked-list method must remain entirely separate and unaffected."""
    campaign_repo = MagicMock()
    campaign_repo.get_by_id.return_value = SimpleNamespace(id=uuid4())
    campaign_candidate_repo = MagicMock()
    campaign_candidate_repo.get_ranked_by_campaign.return_value = ([], 0)
    audit_service = MagicMock()
    service = _make_service(
        campaign_repo=campaign_repo, campaign_candidate_repo=campaign_candidate_repo, audit_service=audit_service,
    )

    service.get_ranked_campaign_candidates(campaign_repo.get_by_id.return_value.id)

    # The read-only list endpoint must never audit (Phase 1's own rule, unchanged).
    audit_service.log.assert_not_called()
