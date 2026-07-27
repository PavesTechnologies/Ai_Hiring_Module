"""
M07-E03 S03 - integration + regression coverage.

Integration: exercises real collaborators end-to-end (the actual
ExcelExport.export_rejected_candidates openpyxl workbook, and the
scorecard + rejection-history flow together for the same candidate across
multiple rejection rounds) instead of mocking every boundary, matching this
suite's existing "integration test" convention
(tests/tasks/test_deterministic_scoring_tasks.py).

Regression: confirms the constructor extension (new optional
candidate_repo/resume_repo/candidate_rejection_repo params) does not break
the one pre-existing call site that constructs this service with only 3
positional args (app/tasks/bulk_upload_tasks.py).
"""

import asyncio
from datetime import datetime, timezone
from types import SimpleNamespace
from unittest.mock import MagicMock
from uuid import uuid4

import openpyxl

from app.enums.constants import ActionType, EntityType
from app.models.pipeline import PipelineStage, RejectionLayer
from app.services.campaign.campaign_candidate_service import CampaignCandidateService


def _drain_streaming_response_body(response) -> bytes:
    """StreamingResponse wraps a sync BytesIO in an async generator (via
    Starlette's iterate_in_threadpool) - drain it the same way an ASGI
    server would, without needing a real HTTP round-trip."""

    async def _collect():
        chunks = []
        async for chunk in response.body_iterator:
            chunks.append(chunk)
        return b"".join(chunks)

    return asyncio.run(_collect())


def _make_campaign_candidate(pipeline_stage=PipelineStage.REJECTED, score_breakdown=None, deterministic_score=35.0):
    return SimpleNamespace(
        id=uuid4(),
        campaign_id=uuid4(),
        candidate_id=uuid4(),
        resume_id=uuid4(),
        pipeline_stage=pipeline_stage,
        score_breakdown=score_breakdown or {},
        hr_override=False,
        hr_override_reason=None,
        deterministic_score=deterministic_score,
        ai_ats_score=None,
        semantic_score=None,
        composite_score=None,
        created_at=datetime.now(timezone.utc),
    )


def _make_rejection(rejection_layer=RejectionLayer.DETERMINISTIC, reason="Missing required skills: Python."):
    return SimpleNamespace(
        id=uuid4(),
        rejection_layer=rejection_layer,
        rejection_reason=reason,
        rejected_at=datetime.now(timezone.utc),
    )


# ----------------------------------------------------------------------
# Integration: real XLSX generation (T03)
# ----------------------------------------------------------------------

def test_export_rejected_candidates_produces_a_real_readable_workbook():
    campaign = SimpleNamespace(id=uuid4())
    breakdown = {
        "mandatory_skills": [
            {"canonical_name": "Python", "match_type": "MISSING"},
            {"canonical_name": "SQL", "match_type": "EXACT"},
        ],
        "experience_validation": {"passed": False, "candidate_years": 1.0, "min_years": 3},
        "education_validation": {"passed": False, "required_level": "BACHELOR", "candidate_level": "HIGH_SCHOOL"},
    }
    candidate = _make_campaign_candidate(score_breakdown=breakdown)
    rejection = _make_rejection(reason="Missing required skills: Python.")

    campaign_repo = MagicMock()
    campaign_repo.get_by_id.return_value = campaign
    campaign_candidate_repo = MagicMock()
    campaign_candidate_repo.get_rejected_by_campaign.return_value = [candidate]
    candidate_rejection_repo = MagicMock()
    candidate_rejection_repo.get_by_campaign_candidate_id.return_value = [rejection]
    audit_service = MagicMock()

    service = CampaignCandidateService(
        campaign_repo=campaign_repo,
        campaign_candidate_repo=campaign_candidate_repo,
        audit_service=audit_service,
        candidate_rejection_repo=candidate_rejection_repo,
    )

    response = service.export_rejected_candidates(campaign.id, actor_id="hr-1", actor_role="HR_ADMIN")

    # Drain the real StreamingResponse body and parse it as a real xlsx file -
    # proves ExcelExport.export_rejected_candidates round-trips through
    # openpyxl correctly, not just that it was called.
    body = _drain_streaming_response_body(response)
    import io
    workbook = openpyxl.load_workbook(io.BytesIO(body))
    sheet = workbook.active

    header = [cell.value for cell in sheet[1]]
    assert header == [
        "Candidate UUID",
        "Rejection Layer",
        "Rejection Reason",
        "Rejected At",
        "Deterministic Score",
        "Missing Mandatory Skills",
        "Experience Gap",
        "Education Gap",
        "HR Override",
        "Override Reason",
    ]

    data_row = [cell.value for cell in sheet[2]]
    assert data_row[0] == str(candidate.candidate_id)
    assert data_row[1] == RejectionLayer.DETERMINISTIC.value
    assert data_row[2] == rejection.rejection_reason
    assert data_row[5] == "Python"
    assert "1.0 years provided" in data_row[6]
    assert "Bachelor" in data_row[7]

    audit_service.log.assert_called_once()


def test_export_rejected_candidates_empty_campaign_still_produces_header_only_workbook():
    campaign = SimpleNamespace(id=uuid4())
    campaign_repo = MagicMock()
    campaign_repo.get_by_id.return_value = campaign
    campaign_candidate_repo = MagicMock()
    campaign_candidate_repo.get_rejected_by_campaign.return_value = []
    audit_service = MagicMock()

    service = CampaignCandidateService(
        campaign_repo=campaign_repo,
        campaign_candidate_repo=campaign_candidate_repo,
        audit_service=audit_service,
    )

    response = service.export_rejected_candidates(campaign.id, actor_id="hr-1", actor_role="HR_ADMIN")

    body = _drain_streaming_response_body(response)
    import io
    workbook = openpyxl.load_workbook(io.BytesIO(body))
    sheet = workbook.active

    assert sheet.max_row == 1  # header only, no data rows
    audit_service.log.assert_called_once_with(
        actor_id="hr-1",
        actor_role="HR_ADMIN",
        action_type=ActionType.REJECTED_CANDIDATES_EXPORTED,
        entity_type=EntityType.CAMPAIGN,
        entity_id=campaign.id,
        campaign_id=campaign.id,
        details={"exported_count": 0},
    )


# ----------------------------------------------------------------------
# Integration: scorecard + rejection history for the same candidate,
# across multiple rejection rounds (T01 + T02 combined)
# ----------------------------------------------------------------------

def test_scorecard_and_history_agree_on_current_rejection_across_multiple_rounds():
    candidate = _make_campaign_candidate(pipeline_stage=PipelineStage.REJECTED)
    round_1 = _make_rejection(reason="Missing required skills: Python.")
    round_2 = _make_rejection(reason="Insufficient experience: 2 years provided, minimum 4 years required (gap: 2 years).")

    campaign_candidate_repo = MagicMock()
    campaign_candidate_repo.get_by_id.return_value = candidate
    candidate_rejection_repo = MagicMock()
    # newest-first, as CandidateRejectionRepository.get_by_campaign_candidate_id guarantees.
    candidate_rejection_repo.get_by_campaign_candidate_id.return_value = [round_2, round_1]

    service = CampaignCandidateService(
        campaign_repo=MagicMock(),
        campaign_candidate_repo=campaign_candidate_repo,
        audit_service=MagicMock(),
        candidate_rejection_repo=candidate_rejection_repo,
    )

    scorecard = service.get_campaign_candidate_scorecard(candidate.id)
    history = service.get_rejection_history(candidate.id)

    # The scorecard banner must reflect the SAME "current" rejection that
    # history marks as current_status=True.
    current_history_entry = next(entry for entry in history if entry.current_status)
    assert scorecard.rejection_reason == current_history_entry.rejection_reason
    assert scorecard.rejected_at == current_history_entry.rejected_at
    assert current_history_entry.id == round_2.id
    assert len(history) == 2
    assert history[0].evaluation_round == 2
    assert history[1].evaluation_round == 1


# ----------------------------------------------------------------------
# Regression: pre-existing call site / methods unaffected
# ----------------------------------------------------------------------

def test_service_still_constructible_with_only_the_original_three_positional_args():
    """
    Matches app/tasks/bulk_upload_tasks.py's exact call:
    CampaignCandidateService(campaign_repo, campaign_candidate_repo, audit_service)
    """
    campaign_repo = MagicMock()
    campaign_candidate_repo = MagicMock()
    audit_service = MagicMock()

    service = CampaignCandidateService(campaign_repo, campaign_candidate_repo, audit_service)

    assert service.candidate_repo is None
    assert service.resume_repo is None
    assert service.candidate_rejection_repo is None
    assert service.encryption_service is None


def test_create_campaign_candidate_unaffected_by_new_optional_repos():
    from app.models.campaigns import CampaignStatus
    from app.schemas.campaign.campaign_candidate_schema import CampaignCandidateCreateRequest

    campaign = SimpleNamespace(id=uuid4(), status=CampaignStatus.ACTIVE, max_candidates=None)
    campaign_repo = MagicMock()
    campaign_repo.get_by_id_for_update.return_value = campaign
    campaign_candidate_repo = MagicMock()
    campaign_candidate_repo.get_by_campaign_and_candidate.return_value = None
    campaign_candidate_repo.get_candidate_count.return_value = 0
    created = _make_campaign_candidate(pipeline_stage=PipelineStage.UPLOADED)
    campaign_candidate_repo.create_idempotent.return_value = (created, True)
    audit_service = MagicMock()

    service = CampaignCandidateService(campaign_repo, campaign_candidate_repo, audit_service)

    request = CampaignCandidateCreateRequest(
        campaign_id=campaign.id,
        candidate_id=created.candidate_id,
        resume_id=created.resume_id,
    )

    result = service.create_campaign_candidate(request, actor_id="user-1", actor_role="RECRUITER")

    assert result.id == created.id
    campaign_candidate_repo.create_idempotent.assert_called_once()


def test_get_campaign_candidates_list_unaffected_by_new_optional_repos():
    campaign_id = uuid4()
    candidate = _make_campaign_candidate(pipeline_stage=PipelineStage.SCREENING)

    campaign_candidate_repo = MagicMock()
    campaign_candidate_repo.get_all_by_campaign.return_value = [(candidate, None, None)]

    service = CampaignCandidateService(MagicMock(), campaign_candidate_repo, MagicMock())

    results = service.get_campaign_candidates(campaign_id)

    assert len(results) == 1
    assert results[0].id == candidate.id
    assert results[0].candidate_name is None
