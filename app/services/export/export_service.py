import logging
from datetime import datetime, timezone
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook

from app.enums.constants import EntityType
from app.exceptions.campaign_exceptions import CampaignException
from app.models.pipeline import DecisionType
from app.repositories.export_repository import ExportRepository
from app.utils.excel_export import ExcelExport
from app.utils.pdf_export import (
    build_pdf, data_table, heading, key_value_table, spacer, title_block,
)

logger = logging.getLogger(__name__)

DEFAULT_EXPORT_ASYNC_THRESHOLD = 500

_CANDIDATE_HEADERS = [
    "Rank", "Candidate ID", "Composite", "Deterministic", "Semantic %",
    "Effective AI", "ATS", "AI Confidence", "AI Recommendation", "Stage",
    "HR Override", "Partial Score", "Days in Stage", "Submitted",
    "Days Since Submission", "Uploaded By", "Recruiter Notes",
]
_CANDIDATE_KEYS = [
    "rank", "candidate_id", "composite", "deterministic", "semantic",
    "effective_ai", "ats", "ai_confidence", "ai_recommendation", "stage",
    "hr_override", "partial_score", "days_in_stage", "submitted",
    "days_since_submission", "uploaded_by", "recruiter_notes",
]


def _v(value):
    """Enum -> value, Decimal -> float, None -> ''. Everything else unchanged."""
    if value is None:
        return ""
    if hasattr(value, "value"):
        return value.value
    if isinstance(value, UUID):
        return str(value)
    try:
        from decimal import Decimal
        if isinstance(value, Decimal):
            return float(value)
    except Exception:
        pass
    return value


def _dt(value):
    if value is None:
        return ""
    return value.strftime("%d-%b-%Y %I:%M %p")


def _days_since(value):
    if value is None:
        return ""
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return (datetime.now(timezone.utc) - value).days


class ExportService:
    """
    Every campaign export.

    Rendering is deliberately split from delivery: each build_* method returns
    bytes and never touches storage, email or the request. That is what lets
    the same builder serve a synchronous download (small exports) and the
    Celery EXPORT_GENERATE path (large ones) without a second implementation
    that can drift from it.
    """

    def __init__(
        self,
        export_repo: ExportRepository,
        campaign_repo,
        audit_service,
        config_repo=None,
        note_repo=None,
    ):
        self.export_repo = export_repo
        self.campaign_repo = campaign_repo
        self.audit_service = audit_service
        self.config_repo = config_repo
        self.note_repo = note_repo

    # ── config ────────────────────────────────────────────────────────

    def _config_int(self, key: str, default: int) -> int:
        if self.config_repo is None:
            return default
        try:
            configs = self.config_repo.get_configs_by_keys([key])
            return int(configs.get(key, default))
        except Exception:
            return default

    def async_threshold(self) -> int:
        return self._config_int("EXPORT_ASYNC_THRESHOLD", DEFAULT_EXPORT_ASYNC_THRESHOLD)

    def _require_campaign(self, campaign_id: UUID):
        campaign = self.campaign_repo.get_by_id(campaign_id)
        if campaign is None:
            raise CampaignException("Campaign not found.", 404)
        return campaign

    def _user_names(self, user_ids) -> dict:
        ids = [u for u in set(user_ids or []) if u]
        if not ids:
            return {}
        try:
            return self.campaign_repo.get_user_names(ids)
        except Exception:
            # A name lookup failure must degrade to raw ids, never fail an export.
            logger.exception("Export user-name lookup failed")
            return {}

    def _notes_map(self, cc_ids) -> dict:
        if self.note_repo is None or not cc_ids:
            return {}
        try:
            out = {}
            for cc_id in cc_ids:
                notes = self.note_repo.list_for_candidate(cc_id)
                if notes:
                    out[str(cc_id)] = " | ".join(n.note_text for n in notes)
            return out
        except Exception:
            # candidate_notes may not exist yet (migration pending) — the rest
            # of the export is still valid without it.
            logger.warning("Recruiter notes unavailable for export", exc_info=True)
            return {}

    # ── candidate list XLSX ─────────────────────────────

    def candidate_list_row_count(
        self, campaign_id: UUID, campaign_candidate_ids: list[UUID] | None = None,
    ) -> int:
        return len(self.export_repo.candidates_for_export(
            campaign_id, campaign_candidate_ids=campaign_candidate_ids,
        ))

    def build_candidate_list_xlsx(
        self,
        campaign_id: UUID,
        *,
        campaign_candidate_ids: list[UUID] | None = None,
        include_rejected_sheet: bool = False,
    ) -> bytes:
        campaign = self._require_campaign(campaign_id)
        rows = self.export_repo.candidates_for_export(
            campaign_id, campaign_candidate_ids=campaign_candidate_ids,
        )
        stage_days = self.export_repo.days_in_current_stage(campaign_id)
        names = self._user_names([r.uploaded_by for r in rows])
        notes = self._notes_map([r.campaign_candidate_id for r in rows])

        wb = Workbook()
        ws = wb.active
        ws.title = "Ranked Candidates"

        data = []
        for idx, r in enumerate(rows, start=1):
            cc_id = str(r.campaign_candidate_id)
            data.append({
                "rank": idx,
                # No PII anywhere on this sheet — candidate UUID only.
                "candidate_id": str(r.candidate_id),
                "composite": _v(r.composite_score),
                "deterministic": _v(r.deterministic_score),
                # semantic_score is stored 0–1; the column is a percentage.
                "semantic": round(float(r.semantic_score) * 100, 2) if r.semantic_score is not None else "",
                "effective_ai": _v(r.effective_ai_score),
                "ats": _v(r.ai_ats_score),
                "ai_confidence": _v(r.ai_confidence),
                "ai_recommendation": _v(r.ai_recommendation),
                "stage": _v(r.pipeline_stage),
                "hr_override": "YES" if r.decision_type == DecisionType.RESET else "NO",
                # "partial" = scored without a completed AI evaluation, so the
                # composite rests on fewer layers than a full one.
                "partial_score": "YES" if (
                    r.composite_score is not None and r.ai_ats_score is None
                ) else "NO",
                "days_in_stage": stage_days.get(cc_id, ""),
                "submitted": _dt(r.created_at),
                "days_since_submission": _days_since(r.created_at),
                "uploaded_by": names.get(r.uploaded_by, r.uploaded_by or ""),
                "recruiter_notes": notes.get(cc_id, ""),
            })

        ExcelExport._write_sheet(ws, _CANDIDATE_HEADERS, data, _CANDIDATE_KEYS)

        if include_rejected_sheet:
            self._append_rejection_sheet(wb, campaign_id)

        # Summary sheet last so the workbook always carries its own provenance.
        self._append_context_sheet(wb, campaign, len(data))

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _append_rejection_sheet(self, wb: Workbook, campaign_id: UUID) -> None:
        """Every rejection event, not just the current one."""
        rows = self.export_repo.rejection_rows(campaign_id)
        ws = wb.create_sheet("Rejected Candidates")
        headers = [
            "Candidate ID", "Attempt", "Rejection Layer", "Rejection Reason",
            "Rejected At", "Source", "HR Override", "Missing Skills",
            "Experience Gap", "Education Gap",
        ]
        keys = [
            "candidate_id", "attempt", "layer", "reason", "rejected_at",
            "source", "hr_override", "missing_skills", "experience_gap", "education_gap",
        ]
        data = []
        attempts: dict[str, int] = {}
        for r in rows:
            cid = str(r.candidate_id)
            attempts[cid] = attempts.get(cid, 0) + 1
            details = r.decision_details or {}
            # A candidate whose override was later cleared has the original
            # decision under a different key — check both shapes.
            missing = (
                details.get("missing_mandatory_skills")
                or details.get("missing_skills")
                or (details.get("deterministic_breakdown") or {}).get("missing_mandatory_skills")
                or []
            )
            data.append({
                "candidate_id": cid,
                "attempt": attempts[cid],
                "layer": _v(r.decision_source),
                "reason": r.change_reason or r.decision_reason or "",
                "rejected_at": _dt(r.rejected_at),
                "source": _v(r.transition_source),
                "hr_override": "YES" if r.decision_type == DecisionType.RESET else "NO",
                "missing_skills": ", ".join(str(m) for m in missing) if isinstance(missing, list) else str(missing),
                "experience_gap": str(details.get("experience_gap", "")),
                "education_gap": str(details.get("education_gap", "")),
            })
        ExcelExport._write_sheet(ws, headers, data, keys)

    def _append_context_sheet(self, wb: Workbook, campaign, row_count: int) -> None:
        ws = wb.create_sheet("Export Context")
        headers = ["Field", "Value"]
        keys = ["field", "value"]
        data = [
            {"field": "Campaign", "value": campaign.name},
            {"field": "Campaign status", "value": _v(campaign.status)},
            {"field": "Rows exported", "value": row_count},
            {"field": "Generated at", "value": _dt(datetime.now(timezone.utc))},
            {"field": "Weight — deterministic", "value": _v(campaign.weight_deterministic)},
            {"field": "Weight — semantic", "value": _v(campaign.weight_semantic)},
            {"field": "Weight — AI", "value": _v(campaign.weight_ai)},
            {"field": "PII included", "value": "NO — candidate UUID only"},
        ]
        ExcelExport._write_sheet(ws, headers, data, keys)

    # ── single scorecard PDF ──────────────────────────────────

    def _scorecard_flowables(self, campaign, row) -> list:
        resume = self.export_repo.resume_for(row.resume_id)
        skills = self.export_repo.scorecard_skills(row.resume_id)
        history = self.export_repo.stage_history_for(row.campaign_candidate_id)

        out = []
        out += title_block(
            "Candidate Scorecard",
            f"Candidate {row.candidate_id} · Campaign: {campaign.name}",
        )

        out.append(heading("Score Summary"))
        out.append(key_value_table([
            ("Composite score", _v(row.composite_score)),
            ("Deterministic", f"{_v(row.deterministic_score)} "
                              f"({'passed' if row.deterministic_passed else 'failed'})"),
            ("Semantic", f"{round(float(row.semantic_score) * 100, 2)}%"
                          if row.semantic_score is not None else "—"),
            ("Effective AI score", _v(row.effective_ai_score)),
            ("ATS score", _v(row.ai_ats_score)),
            ("AI confidence", _v(row.ai_confidence)),
            ("AI recommendation", _v(row.ai_recommendation)),
            ("Pipeline stage", _v(row.pipeline_stage)),
            ("Prompt version", _v(row.prompt_version_id)),
        ]))

        out.append(heading("Skill Match"))
        skill_rows = [
            [
                getattr(s, "raw_extracted_text", "") or "",
                _v(getattr(s, "match_tier", None)),
                _v(getattr(s, "scoring_weight", None)),
            ]
            for s in skills[:60]
        ]
        out.append(data_table(["Skill", "Match Tier", "Weight"], skill_rows))
        if len(skills) > 60:
            out.append(spacer(3))
            out.append(data_table(["Note"], [[f"{len(skills) - 60} further skills omitted."]]))

        out.append(heading("AI Evaluation"))
        out.append(data_table(["Strengths"], [[s] for s in (row.ai_strengths or [])] or [["—"]]))
        out.append(spacer(3))
        out.append(data_table(["Weaknesses"], [[w] for w in (row.ai_weaknesses or [])] or [["—"]]))

        out.append(heading("Resume & Parse Quality"))
        out.append(key_value_table([
            ("Parse status", _v(getattr(resume, "parse_status", None))),
            ("Parse confidence", _v(getattr(resume, "parse_confidence_score", None))),
            ("Parser version", _v(getattr(resume, "parser_version", None))),
            ("Pages", _v(getattr(resume, "page_count", None))),
            ("OCR used", "YES" if getattr(resume, "ocr_used", False) else "NO"),
            ("Skills extracted", len(skills)),
        ]))

        out.append(heading("Decision"))
        out.append(key_value_table([
            ("Decision", _v(row.decision_type)),
            ("Decided by layer", _v(row.decision_source)),
            ("Reason", row.decision_reason or "—"),
            ("Decided at", _dt(row.decision_at)),
            ("HR override in force", "YES" if row.decision_type == DecisionType.RESET else "NO"),
        ]))

        out.append(heading("Activity Timeline"))
        out.append(data_table(
            ["When", "From", "To", "Source", "Reason"],
            [[
                _dt(h.changed_at), _v(h.from_stage), _v(h.to_stage),
                _v(h.transition_source), h.change_reason or "",
            ] for h in history],
        ))
        return out

    def build_scorecard_pdf(self, campaign_id: UUID, campaign_candidate_id: UUID) -> bytes:
        campaign = self._require_campaign(campaign_id)
        rows = self.export_repo.candidates_for_export(
            campaign_id, campaign_candidate_ids=[campaign_candidate_id],
        )
        if not rows:
            raise CampaignException("Candidate not found in this campaign.", 404)
        return build_pdf(self._scorecard_flowables(campaign, rows[0]), title="Candidate Scorecard")

    # ── audit trail XLSX ──────────────────────────────────────

    def build_audit_trail_xlsx(self, campaign_id: UUID) -> bytes:
        campaign = self._require_campaign(campaign_id)
        events = self.export_repo.audit_events(campaign_id)
        transitions = self.export_repo.stage_transitions(campaign_id)

        actor_ids = [e.actor_id for e in events] + [t.changed_by for t in transitions]
        names = self._user_names(actor_ids)

        wb = Workbook()
        ws = wb.active
        ws.title = "All Events"
        ExcelExport._write_sheet(
            ws,
            ["Timestamp", "Action", "Actor", "Actor Role", "Entity Type", "Entity ID",
             "IP Address", "Session ID", "Detail"],
            [{
                "ts": _dt(e.created_at),
                "action": _v(e.action_type),
                "actor": names.get(e.actor_id, e.actor_id or "System"),
                "role": e.actor_role or "",
                "etype": _v(e.entity_type),
                "eid": str(e.entity_id),
                "ip": str(e.ip_address) if e.ip_address else "",
                "sid": str(e.session_id) if e.session_id else "",
                "detail": str(e.detail) if e.detail else "",
            } for e in events],
            ["ts", "action", "actor", "role", "etype", "eid", "ip", "sid", "detail"],
        )

        ws2 = wb.create_sheet("Stage Transitions")
        ExcelExport._write_sheet(
            ws2,
            ["Candidate ID", "From", "To", "Changed By", "Source", "Reason", "Changed At"],
            [{
                "cid": str(t.candidate_id),
                "from": _v(t.from_stage),
                "to": _v(t.to_stage),
                # An unattributed transition is the system acting, not a gap.
                "by": names.get(t.changed_by, t.changed_by) if t.changed_by else "System",
                "src": _v(t.transition_source),
                "reason": t.change_reason or "",
                "at": _dt(t.changed_at),
            } for t in transitions],
            ["cid", "from", "to", "by", "src", "reason", "at"],
        )

        ws3 = wb.create_sheet("Score History")
        score_actions = {
            "DETERMINISTIC_SCORE_COMPUTED", "SEMANTIC_SCORE_COMPUTED",
            "COMPOSITE_SCORE_COMPUTED", "AI_EVALUATION_COMPLETED",
        }
        ExcelExport._write_sheet(
            ws3,
            ["Timestamp", "Action", "Entity ID", "Detail"],
            [{
                "ts": _dt(e.created_at),
                "action": _v(e.action_type),
                "eid": str(e.entity_id),
                "detail": str(e.detail) if e.detail else "",
            } for e in events if _v(e.action_type) in score_actions],
            ["ts", "action", "eid", "detail"],
        )

        self._append_context_sheet(wb, campaign, len(events))
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── audit ─────────────────────────────────────────────────────────

    def log_export(self, *, actor_id, actor_role, campaign_id, action_type, details) -> None:
        """Every export is itself an auditable event (all four S0x stories)."""
        try:
            self.audit_service.log(
                actor_id=actor_id,
                actor_role=actor_role,
                action_type=action_type,
                entity_type=EntityType.CAMPAIGN.value,
                entity_id=campaign_id,
                campaign_id=campaign_id,
                details=details,
            )
            self.export_repo.db.commit()
        except Exception:
            self.export_repo.db.rollback()
            # Never fail a completed export because its audit row failed.
            logger.exception("Failed to write export audit entry")
