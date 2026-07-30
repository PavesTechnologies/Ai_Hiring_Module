from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models import jd
class ExcelExport:

    HEADER_FILL = PatternFill(
        fill_type="solid",
        start_color="4F81BD",
    )

    SECTION_FILL = PatternFill(
        fill_type="solid",
        start_color="D9EAD3",
    )

    LABEL_FILL = PatternFill(
        fill_type="solid",
        start_color="F2F2F2",
    )

    HEADER_FONT = Font(
        bold=True,
        color="FFFFFF",
    )

    BOLD_FONT = Font(bold=True)

    ACTIVE_FONT = Font(
        color="008000",
        bold=True,
    )

    CLOSED_FONT = Font(
        color="FF0000",
        bold=True,
    )

    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    @staticmethod
    def export_jd_list(records, user_names):

        wb = Workbook()

        ws = wb.active
        ws.title = "Job Descriptions"

        ws.append([
            "Title",
            "Source Format",
            "Version",
            "Jurisdiction",
            "Experience",
            "Education",
            "Created By",
            "Created At",
            "Status",
            "Linked Campaign Count",
        ])

        header_font = Font(
            bold=True,
            color="FFFFFF",
        )

        header_fill = PatternFill(
            fill_type="solid",
            start_color="4F81BD",
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        # Freeze header
        ws.freeze_panes = "A2"

        for jd in records:

            ws.append([
                jd.title,
                jd.source_format.value,
                jd.version_number,
                jd.jurisdiction,
                float(jd.min_experience_years) if jd.min_experience_years else "",
                jd.education_display,
                user_names.get(jd.created_by, jd.created_by),
                jd.created_at.strftime("%d-%b-%Y %I:%M %p"),
                "Active" if jd.is_active_version else "Closed",
                jd.linked_campaign_count
            ])
        # Auto Filter
        ws.auto_filter.ref = ws.dimensions

        # Auto-fit Columns
        for column_cells in ws.columns:

            max_length = 0

            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )
                except Exception:
                    pass

            ws.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = max_length + 3
        
        for row in ws.iter_rows(min_row=2):

            row[2].alignment = Alignment(horizontal="center")   # Version
            row[4].alignment = Alignment(horizontal="center")   # Experience
            row[7].alignment = Alignment(horizontal="center")   # Created At
            row[8].alignment = Alignment(horizontal="center")   # Status
            row[9].alignment = Alignment(horizontal="center")   # Campaign Count

        output = BytesIO()

        wb.save(output)

        output.seek(0)

        return output
    

    @staticmethod
    def export_skill_ontology(records):
        """records: iterable of (SkillOntology, parent_canonical_name | None) tuples."""

        wb = Workbook()
        ws = wb.active
        ws.title = "Skill Ontology"

        ws.append([
            "Canonical Name",
            "Aliases",
            "Category",
            "Parent Skill",
            "Confidence",
            "Source",
            "Status",
            "Occurrences",
            "Created At",
        ])

        for cell in ws[1]:
            cell.font = ExcelExport.HEADER_FONT
            cell.fill = ExcelExport.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

        for skill, parent_name in records:
            ws.append([
                skill.canonical_name,
                ", ".join(skill.aliases or []),
                skill.category or "",
                parent_name or "",
                skill.confidence,
                skill.source or "",
                "Active" if skill.is_active else "Inactive",
                skill.occurrence_count,
                skill.created_at.strftime("%d-%b-%Y %I:%M %p"),
            ])

        ws.auto_filter.ref = ws.dimensions

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = min(max_length + 3, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def export_rejected_candidates(rows):
        """
        M07-E03 S03 T03: rows is an iterable of dicts (see
        CampaignCandidateService._to_export_row) - one row per rejected
        candidate. Never includes candidate name/email/phone/resume or any
        other PII - only the opaque candidate_uuid plus rejection/score
        fields already computed by the deterministic scoring pipeline.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Rejected Candidates"

        ws.append([
            "Candidate UUID",
            "Rejection Layer",
            "Rejection Reason",
            "Rejected At",
            "Deterministic Score",
            "Missing Mandatory Skills",
            "Experience Gap",
            "Education Gap",
            "HR Override",
            "Override Reason",
        ])

        for cell in ws[1]:
            cell.font = ExcelExport.HEADER_FONT
            cell.fill = ExcelExport.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

        for row in rows:
            ws.append([
                row["candidate_uuid"],
                row["rejection_layer"],
                row["rejection_reason"],
                row["rejected_at"].strftime("%d-%b-%Y %I:%M %p") if row["rejected_at"] else "",
                row["deterministic_score"] if row["deterministic_score"] is not None else "",
                row["missing_mandatory_skills"],
                row["experience_gap"],
                row["education_gap"],
                "Yes" if row["hr_override"] else "No",
                row["override_reason"],
            ])

        ws.auto_filter.ref = ws.dimensions

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = min(max_length + 3, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def export_override_report(rows):
        """
        M07-E03 S04 T03: rows is an iterable of dicts (see
        CampaignCandidateService._to_override_export_row) - one row per HR
        override event. Never includes candidate name/email/phone/resume -
        only the opaque candidate_uuid. HR full name is not candidate PII
        and is included as explicitly required by this report.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Override Report"

        ws.append([
            "Campaign Name",
            "Candidate UUID",
            "Original Rejection Reason",
            "Override Reason",
            "HR Full Name",
            "Override Timestamp",
            "Current Pipeline Stage",
        ])

        for cell in ws[1]:
            cell.font = ExcelExport.HEADER_FONT
            cell.fill = ExcelExport.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

        for row in rows:
            ws.append([
                row["campaign_name"],
                row["candidate_uuid"],
                row["original_rejection_reason"],
                row["override_reason"],
                row["hr_full_name"],
                row["override_timestamp"].strftime("%d-%b-%Y %I:%M %p") if row["override_timestamp"] else "",
                row["current_pipeline_stage"],
            ])

        ws.auto_filter.ref = ws.dimensions

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = min(max_length + 3, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def export_bulk_import_errors(failures):
        """
        S07-T03: failures is an iterable of dicts with the original uploaded
        columns (canonical_name, aliases, category, parent_skill, confidence)
        plus a reason — one row per failed record from a bulk import run, so
        HR_ADMIN can correct and re-upload via the existing import endpoint.
        """

        wb = Workbook()
        ws = wb.active
        ws.title = "Import Errors"

        ws.append([
            "Canonical Name",
            "Aliases",
            "Category",
            "Parent Skill",
            "Confidence",
            "Reason",
        ])

        for cell in ws[1]:
            cell.font = ExcelExport.HEADER_FONT
            cell.fill = ExcelExport.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

        for failure in failures:
            aliases = failure.get("aliases") or []
            ws.append([
                failure.get("canonical_name") or "",
                ", ".join(aliases) if isinstance(aliases, list) else str(aliases),
                failure.get("category") or "",
                failure.get("parent_skill") or "",
                failure.get("confidence") or "",
                failure.get("reason") or "",
            ])

        ws.auto_filter.ref = ws.dimensions

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = min(max_length + 3, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def export_bulk_upload_history(records):
        """
        M05-E02 Phase B8: one row per bulk_upload_jobs record for a
        campaign — a job-level summary, not a per-file breakdown (that's
        what the detail endpoint is for).
        """

        wb = Workbook()
        ws = wb.active
        ws.title = "Bulk Upload History"

        ws.append([
            "File Name",
            "Status",
            "Total Files",
            "Processed",
            "Failed",
            "Duplicate",
            "Uploaded By",
            "Created At",
            "Completed At",
        ])

        for cell in ws[1]:
            cell.font = ExcelExport.HEADER_FONT
            cell.fill = ExcelExport.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

        for job in records:
            ws.append([
                job.original_filename,
                job.status.value,
                job.total_files,
                job.processed_count,
                job.failed_count,
                job.duplicate_count,
                job.uploaded_by,
                job.created_at.strftime("%d-%b-%Y %I:%M %p"),
                job.completed_at.strftime("%d-%b-%Y %I:%M %p") if job.completed_at else "",
            ])

        ws.auto_filter.ref = ws.dimensions

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = min(max_length + 3, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def export_single_jd(
        jd,
        version_history,
        created_by_name,
        linked_campaigns,
    ):

        wb = Workbook()
        ws = wb.active
        ws.title = "Job Description"
        ws.freeze_panes = "A2"
        bold = Font(bold=True)

        section_font = Font(
            bold=True,
            size=13,
        )

        section_fill = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
        )

        label_fill = PatternFill(
            fill_type="solid",
            start_color="F2F2F2",
        )

        header_fill = PatternFill(
            fill_type="solid",
            start_color="4F81BD",
        )

        header_font = Font(
            bold=True,
            color="FFFFFF",
        )

        row = 1

        # ======================
        # JD DETAILS
        # ======================

        ws.merge_cells(
                start_row=row,
                start_column=1,
                end_row=row,
                end_column=2,
            )

        cell = ws.cell(row=row, column=1)
        cell.value = "JOB DESCRIPTION DETAILS"
        cell.font = section_font
        cell.fill = section_fill
        row += 2

        details = [
            ("Title", jd.title),
            ("Source Format", jd.source_format.value),
            ("Version", jd.version_number),
            ("Jurisdiction", jd.jurisdiction),
            ("Minimum Experience", jd.min_experience_years),
            ("Created By", created_by_name),
            ("Created At", jd.created_at.strftime("%d-%b-%Y %I:%M %p")),
            ("Status", "Active" if jd.is_active_version else "Closed"),
        ]

        for key, value in details:
            label = ws.cell(row=row, column=1)

            label.value = key
            label.font = bold
            label.fill = label_fill
            ws.cell(row=row, column=2).value = str(value) if value else ""
            label.border = ExcelExport.THIN_BORDER
            ws.cell(row=row, column=2).border = ExcelExport.THIN_BORDER
            row += 1

        row += 2

        # ======================
        # EDUCATION
        # ======================

        cell = ws.cell(row=row, column=1)
        cell.value = "EDUCATION CRITERIA"
        cell.font = section_font
        cell.fill = section_fill
        row += 1

        if jd.education_criteria:
            degree = jd.education_criteria.get("degree", "")
            field = jd.education_criteria.get("field", "")

            label = ws.cell(row=row, column=1)
            label.value = "Degree"
            label.font = bold
            label.fill = label_fill

            ws.cell(row=row, column=2).value = degree

            row += 1

            label = ws.cell(row=row, column=1)
            label.value = "Field"
            label.font = bold
            label.fill = label_fill

            ws.cell(row=row, column=2).value = field
            row += 1
        else:
            ws.cell(row=row, column=1).value = "N/A"
            row += 1

        row += 2

        # ======================
        # RAW TEXT
        # ======================

        cell = ws.cell(row=row, column=1)
        cell.value = "Raw Text"
        cell.font = section_font
        cell.fill = section_fill
        row += 1

        cell = ws.cell(
            row=row,
            column=1,
        )

        cell.value = jd.raw_text or ""


        cell.alignment = Alignment(
            wrap_text=True,
            vertical="top",
    )

        ws.row_dimensions[row].height = 90

        row += 3

        # ======================
        # PARSED SKILLS
        # ======================

        cell = ws.cell(row=row, column=1)
        cell.value = "PARSED SKILLS"
        cell.font = section_font
        cell.fill = section_fill
        row += 1

        if jd.extracted_json:

            if isinstance(jd.extracted_json, dict):

                for key, value in jd.extracted_json.items():

                    ws.cell(row=row, column=1).value = key
                    ws.cell(row=row, column=2).value = str(value)

                    row += 1

            elif isinstance(jd.extracted_json, list):

                for skill in jd.extracted_json:

                    ws.cell(row=row, column=1).value = str(skill)

                    row += 1

        else:

            ws.cell(row=row, column=1).value = "N/A"

            row += 1

        row += 2

        # ======================
        # REQUIRED SKILLS
        # ======================

        cell = ws.cell(row=row, column=1)
        cell.value = "Required Skills"
        cell.font = section_font
        cell.fill = section_fill
        row += 1

        if jd.required_skills:

            if isinstance(jd.required_skills, dict):

                for key, value in jd.required_skills.items():

                    ws.cell(row=row, column=1).value = key
                    ws.cell(row=row, column=2).value = str(value)

                    row += 1

            elif isinstance(jd.required_skills, list):

                for skill in jd.required_skills:

                    ws.cell(row=row, column=1).value = str(skill)

                    row += 1

        else:

            ws.cell(row=row, column=1).value = "N/A"

            row += 1

        row += 2

        # ======================
        # VERSION HISTORY
        # ======================

        cell = ws.cell(row=row, column=1)
        cell.value = "Version History"
        cell.font = section_font
        cell.fill = section_fill

        row += 1

        ws.append([
            "Version",
            "Created At",
            "Status"
        ])

        header_row = ws.max_row

        for col in range(1, 4):

            cell = ws.cell(
                row=header_row,
                column=col,
            )

            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for version in version_history:

            ws.append([
                version.version_number,
                version.created_at.strftime("%Y-%m-%d %H:%M"),
                "Active" if version.is_active_version else "Closed"
            ])

            current = ws.max_row

            for c in ws[current]:
                c.border = ExcelExport.THIN_BORDER

            ws.cell(current,1).alignment = Alignment(horizontal="center")
            ws.cell(current,2).alignment = Alignment(horizontal="center")

            status = ws.cell(current,3)

            status.alignment = Alignment(horizontal="center")

            status.font = (
                ExcelExport.ACTIVE_FONT
                if version.is_active_version
                else ExcelExport.CLOSED_FONT
            )

        # ======================
        # LINKED CAMPAIGNS
        # ======================

        row += 3

        cell = ws.cell(row=row, column=1)
        cell.value = "Linked Campaigns"
        cell.font = section_font
        cell.fill = section_fill

        row += 1

        campaign_header = row
        ws.cell(row=row, column=1).value = "Campaign Name"
        ws.cell(row=row, column=2).value = "Status"
        for col in range(1, 3):

            cell = ws.cell(
                row=campaign_header,
                column=col,
            )

            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row, column=1).font = bold
        ws.cell(row=row, column=2).font = bold

        row += 1

        if linked_campaigns:

            for campaign in linked_campaigns:

                ws.cell(row=row, column=1).value = campaign.name
                ws.cell(row=row, column=2).value = campaign.status.value

                for c in ws[row]:
                    c.border = ExcelExport.THIN_BORDER

                ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

                row += 1
        else:
                ws.cell(row=row, column=1).value = "No linked campaigns"


        for column_cells in ws.columns:

            max_length = 0

            for cell in column_cells:

                try:

                    if cell.value:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                except Exception:
                    pass

            ws.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = min(max_length + 3, 50)

        output = BytesIO()

        wb.save(output)

        output.seek(0)

        return output

    @staticmethod
    def _write_sheet(ws, headers, rows, row_key_order):
        """
        M07-E03 S05 T03: shared per-sheet writer for
        export_deterministic_rejection_summary - same header-styling/
        freeze-panes/auto-filter/auto-fit conventions as every other
        single-sheet export method in this file, just parameterized so
        one workbook can hold 3 sheets without repeating the boilerplate
        3 times.
        """
        ws.append(headers)

        for cell in ws[1]:
            cell.font = ExcelExport.HEADER_FONT
            cell.fill = ExcelExport.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

        for row in rows:
            ws.append([
                # openpyxl rejects tz-aware datetimes outright - format the
                # same way every other export method in this file already
                # does, rather than making every S05 row-builder remember to.
                value.strftime("%d-%b-%Y %I:%M %p") if isinstance(value, datetime) else value
                for value in (row.get(key, "") for key in row_key_order)
            ])

        ws.auto_filter.ref = ws.dimensions

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = min(max_length + 3, 50)

    @staticmethod
    def export_deterministic_rejection_summary(campaign_summary_rows, skill_gap_rows, override_log_rows):
        """
        M07-E03 S05 T03: platform-wide deterministic rejection summary -
        3 sheets in one workbook (Campaign Summary, Skill Gap Analysis,
        Override Log). Never includes candidate name/email/phone/resume -
        only the opaque candidate_uuid on the Override Log sheet.
        """
        wb = Workbook()
        wb.remove(wb.active)

        ExcelExport._write_sheet(
            wb.create_sheet("Campaign Summary"),
            [
                "Campaign Name",
                "Total Candidates",
                "Deterministic Rejections",
                "Rejection Rate (%)",
                "Top Rejection Reason",
                "Override Count",
                "Override Rate (%)",
            ],
            campaign_summary_rows,
            [
                "campaign_name",
                "total_candidates",
                "deterministic_rejections",
                "rejection_rate",
                "top_rejection_reason",
                "override_count",
                "override_rate",
            ],
        )

        ExcelExport._write_sheet(
            wb.create_sheet("Skill Gap Analysis"),
            [
                "Skill Canonical Name",
                "Campaigns Requiring Skill",
                "Missing Count",
                "Missing Rate (%)",
            ],
            skill_gap_rows,
            [
                "canonical_name",
                "campaigns_requiring_skill",
                "missing_count",
                "missing_rate",
            ],
        )

        ExcelExport._write_sheet(
            wb.create_sheet("Override Log"),
            [
                "Campaign Name",
                "Candidate UUID",
                "Rejection Reason",
                "Override Reason",
                "Override By",
                "Override Timestamp",
                "Current Pipeline Stage",
            ],
            override_log_rows,
            [
                "campaign_name",
                "candidate_uuid",
                "rejection_reason",
                "override_reason",
                "override_by",
                "override_timestamp",
                "current_pipeline_stage",
            ],
        )

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output