from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REQUIRED_COLUMNS = [
    "canonical_name",
    "aliases",
    "category",
    "parent_skill",
    "confidence",
    "source",
    "is_active",
]


class SkillExcelReader:
    """
    Parses Skill Ontology rows out of an Excel workbook.

    Pure parsing utility: performs no database access, so it can be shared
    as-is between the seed script and the future Bulk Import feature.
    """

    @staticmethod
    def read(file_path: str | Path) -> list[dict[str, Any]]:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Skill ontology Excel file not found: {path}")

        # read_only workbooks keep the underlying file handle open until
        # explicitly closed — without this, deleting the (often temporary)
        # source file right after read() fails on Windows with a
        # PermissionError, since the handle is still held.
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active

            rows = worksheet.iter_rows(values_only=True)

            # The header is normally row 1, but some tools (Google Sheets
            # exports, manually-built workbooks) prepend a title row or
            # blank spacer row above it. Scan the first few rows for one
            # that actually contains the required columns instead of only
            # ever looking at the very first row.
            header_row = None
            headers: list[str] = []
            first_row_headers: list[str] | None = None
            header_row_number = 0
            for header_row_number in range(1, 6):
                candidate = next(rows, None)
                if candidate is None:
                    break
                candidate_headers = [SkillExcelReader._normalize_header(cell) for cell in candidate]
                if first_row_headers is None:
                    first_row_headers = candidate_headers
                if all(column in candidate_headers for column in REQUIRED_COLUMNS):
                    header_row = candidate
                    headers = candidate_headers
                    break

            if header_row is None:
                if first_row_headers is None:
                    raise ValueError(f"Skill ontology Excel file has no header row: {path}")
                missing_columns = [column for column in REQUIRED_COLUMNS if column not in first_row_headers]
                raise ValueError(
                    f"Skill ontology Excel is missing required column(s): {', '.join(missing_columns)}. "
                    f"Detected header row: {first_row_headers}"
                )

            column_index = {column: headers.index(column) for column in REQUIRED_COLUMNS}

            skills: list[dict[str, Any]] = []
            for row_number, row in enumerate(rows, start=header_row_number + 1):
                if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
                    continue  # skip completely blank rows (spreadsheet padding, not real data)

                skill = SkillExcelReader._parse_row(row, column_index)
                skill["row_number"] = row_number
                # A blank canonical_name (or a blank alias entry) is intentionally
                # NOT filtered out here — this reader only parses/normalizes.
                # Deciding whether that makes a row valid is the caller's job
                # (bulk-import validation reports it; execution treats it as a
                # failed row) so both flows see the same row numbering.
                skills.append(skill)

            return skills
        finally:
            workbook.close()

    @staticmethod
    def _parse_row(row: tuple, column_index: dict[str, int]) -> dict[str, Any]:
        def cell(column: str) -> Any:
            index = column_index[column]
            return row[index] if index < len(row) else None

        def clean_str(value: Any) -> str:
            return str(value).strip() if value is not None else ""

        aliases_raw = clean_str(cell("aliases"))
        aliases = [alias.strip() for alias in aliases_raw.split(",")] if aliases_raw else []

        return {
            "canonical_name": clean_str(cell("canonical_name")),
            "aliases": aliases,
            "category": clean_str(cell("category")) or None,
            "parent_skill": clean_str(cell("parent_skill")) or None,
            "confidence": clean_str(cell("confidence")).lower() or "unverified",
            "source": clean_str(cell("source")).lower() or None,
            "is_active": SkillExcelReader._parse_bool(cell("is_active")),
        }

    @staticmethod
    def _parse_bool(value: Any) -> bool:
        if isinstance(value, bool):
            return value
        return str(value).strip().upper() in ("TRUE", "ACTIVE", "YES", "1")

    @staticmethod
    def _normalize_header(cell: Any) -> str:
        if cell is None:
            return ""
        # Headers pasted from Word/PDF/web pages often carry a BOM or
        # non-breaking space instead of a plain one, which plain str.strip()
        # leaves behind and makes an otherwise-correct header fail the
        # REQUIRED_COLUMNS membership check.
        return str(cell).replace("﻿", "").replace("\xa0", " ").strip().lower()
